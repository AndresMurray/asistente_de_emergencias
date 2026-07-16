"""
Evaluación completa del RAG a partir de "DataSet Test.xlsx".

Calcula, para cada pregunta del Excel:
  - STS (Semantic Text Similarity) vs. la respuesta esperada
  - Critical Information Coverage
  - Answer Relevancy (estilo RAGAS)
  - MRR@10 (retrieval)
  - Recall@5 (retrieval)

Y reporta promedios globales.

Uso:
    python metricas/evaluar_excel_completo.py [--modelo MODELO]

Requisitos:
  - Stack levantado: PostgreSQL/pgvector en :5433 con corpus ingestado + Ollama.
  - Archivos "DataSet Test.xlsx" y "metricas/dataset_evaluacion.json" en sus lugares.
"""

import argparse
import asyncio
import difflib
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import requests

# Asegurar que src/ y metricas/ sean importables desde cualquier cwd
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from metricas.answer_relevancy import (
    N_QUESTIONS,
    build_reverse_question_prompt,
    cosine_similarity,
    is_noncommittal,
    parse_questions,
)
from metricas.critical_information_coverage import critical_information_coverage
from metricas.mrr import fetch_corpus, gold_ids_for_query, reciprocal_rank
from metricas.recall_at_5 import recall_at_k
from metricas.sts import sts
from src.main import Pipeline

# ── Configuración ─────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(ROOT, "DataSet Test.xlsx")
DATASET_JSON_PATH = os.path.join(ROOT, "metricas", "dataset_evaluacion.json")
OUTPUT_DIR = os.path.join(ROOT, "metricas", "resultados")
SHEET_NAME = "Hoja 1"
K_MRR = 10
K_RECALL = 5
MIN_HITS = 1

os.environ.setdefault(
    "DATABASE_URL",
    "postgresql://postgres:postgres@localhost:5433/emergencias_vdb",
)


def collect_stream(stream) -> str:
    """El pipeline devuelve un generador o un string; lo unificamos a texto."""
    if isinstance(stream, str):
        return stream
    return "".join(token for token in stream)


def load_excel(path: str) -> List[Dict[str, Any]]:
    """Lee el Excel y devuelve lista de entradas {fila, query, expected_answer}."""
    if not Path(path).exists():
        raise FileNotFoundError(f"No se encontró el Excel: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]

    rows = []
    for num_fila, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        query = row[0]
        expected = row[1]
        if query is None or not str(query).strip():
            continue
        rows.append(
            {
                "fila": num_fila,
                "query": str(query).strip(),
                "expected_answer": str(expected).strip() if expected else "",
            }
        )
    wb.close()
    return rows


def normalize_for_match(text: str) -> str:
    """Normaliza una pregunta para compararla contra el dataset JSON."""
    return " ".join(text.lower().split())


def match_queries_to_dataset(
    excel_rows: List[Dict[str, Any]], dataset: List[Dict[str, Any]]
) -> Dict[str, Optional[Dict[str, Any]]]:
    """Mapea cada query del Excel a su entrada en dataset_evaluacion.json."""
    json_by_query = {normalize_for_match(d["query"]): d for d in dataset}
    mapping: Dict[str, Optional[Dict[str, Any]]] = {}

    for row in excel_rows:
        q_norm = normalize_for_match(row["query"])
        match = json_by_query.get(q_norm)

        if match is None:
            # Fallback por similitud de secuencia (por si hay una frase distinta)
            best = difflib.get_close_matches(
                q_norm, json_by_query.keys(), n=1, cutoff=0.75
            )
            if best:
                match = json_by_query[best[0]]

        mapping[row["query"]] = match
    return mapping


def calculate_answer_relevancy(
    pipe: Pipeline, query: str, answer: str
) -> Tuple[float, List[str], List[float]]:
    """Devuelve (score, preguntas_inversas, similitudes)."""
    if is_noncommittal(answer):
        return 0.0, [], []

    prompt = build_reverse_question_prompt(answer, N_QUESTIONS)
    raw_questions = collect_stream(pipe.llm_client.generate_stream(prompt))
    gen_questions = parse_questions(raw_questions, N_QUESTIONS)

    if not gen_questions:
        return 0.0, [], []

    q_emb = pipe.searcher.get_embeddings(query)
    sims = [cosine_similarity(q_emb, pipe.searcher.get_embeddings(gq)) for gq in gen_questions]
    score = sum(sims) / len(sims)
    return score, gen_questions, sims


def ollama_model_exists(model_name: str) -> bool:
    """Verifica si el modelo está disponible en Ollama local."""
    try:
        base_url = os.environ.get("OLLAMA_URL", "http://localhost:11434")
        resp = requests.get(f"{base_url}/api/tags", timeout=10)
        resp.raise_for_status()
        models = resp.json().get("models", [])
        return any(m.get("name", m.get("model")) == model_name for m in models)
    except Exception as e:
        print(f"[Aviso] No se pudo verificar modelos en Ollama: {e}")
        return True  # No bloquear si Ollama no responde todavía


async def main():
    parser = argparse.ArgumentParser(
        description="Evaluación completa del RAG usando el Excel de preguntas."
    )
    parser.add_argument(
        "--modelo",
        default=os.environ.get("MODEL_NAME", "gemma2:2b"),
        help="Modelo de Ollama a usar (default: gemma2:2b o $MODEL_NAME).",
    )
    parser.add_argument(
        "--excel",
        default=EXCEL_PATH,
        help=f"Ruta al Excel de preguntas (default: {EXCEL_PATH}).",
    )
    parser.add_argument(
        "--dataset",
        default=DATASET_JSON_PATH,
        help=f"Ruta al dataset JSON con critical_facts/keywords (default: {DATASET_JSON_PATH}).",
    )
    args = parser.parse_args()

    model_name = args.modelo
    excel_path = args.excel
    dataset_path = args.dataset

    print("=" * 80)
    print(" EVALUACIÓN COMPLETA DEL RAG DESDE EXCEL")
    print("=" * 80)
    print(f" Excel: {excel_path}")
    print(f" Dataset: {dataset_path}")
    print(f" Modelo: {model_name}")
    print(f" DB: {os.environ['DATABASE_URL']}")
    print("=" * 80)

    if not ollama_model_exists(model_name):
        print(f"\n[ERROR] El modelo '{model_name}' no está disponible en Ollama.")
        print("  Descargalo con: ollama pull {model_name}")
        sys.exit(1)

    # 1. Cargar Excel y JSON
    excel_rows = load_excel(excel_path)
    with open(dataset_path, "r", encoding="utf-8") as f:
        dataset = json.load(f)

    query_to_dataset = match_queries_to_dataset(excel_rows, dataset)

    unmatched = [r["query"] for r in excel_rows if query_to_dataset[r["query"]] is None]
    if unmatched:
        print(f"\n[ADVERTENCIA] {len(unmatched)} pregunta(s) del Excel no se cruzaron con el JSON:")
        for q in unmatched:
            print(f"  - {q}")
        print()

    # 2. Inicializar pipeline con el modelo elegido
    print("Inicializando pipeline local...")
    pipe = Pipeline()
    pipe.valves.MODEL_NAME = model_name
    await pipe.on_startup()

    # 3. Cargar corpus completo para retrieval (MRR/Recall)
    corpus = fetch_corpus(pipe.searcher.db_manager)
    if corpus:
        print(f"Corpus cargado: {len(corpus)} chunks.\n")
    else:
        print("[ADVERTENCIA] No se pudo cargar el corpus completo. MRR/Recall pueden degradarse.\n")

    # 4. Evaluar cada pregunta
    results: List[Dict[str, Any]] = []
    totals = {
        "sts": 0.0,
        "critical_info_coverage": 0.0,
        "answer_relevancy": 0.0,
        "mrr": 0.0,
        "recall_at_5": 0.0,
    }
    counts = {
        "sts": 0,
        "critical_info_coverage": 0,
        "answer_relevancy": 0,
        "mrr": 0,
        "recall_at_5": 0,
    }

    print(f"{'ID':<4} | {'Query':<45} | {'STS':>6} | {'CIC':>5} | {'Rel':>5} | {'MRR':>5} | {'R@5':>5}")
    print("-" * 80)

    for idx, row in enumerate(excel_rows, start=1):
        query = row["query"]
        expected = row["expected_answer"]
        ds_entry = query_to_dataset[query]

        t0 = time.perf_counter()

        # Generar respuesta real del RAG
        generated = collect_stream(pipe.pipe(query, pipe.valves.MODEL_NAME, [])).strip()

        # STS vs. respuesta esperada
        sts_score = None
        if expected:
            try:
                sts_score = sts(generated, expected)
            except Exception as e:
                print(f"[STS error] {e}")

        # Critical Information Coverage
        cic_score = None
        if ds_entry:
            critical_facts = ds_entry.get("critical_facts", [])
            if critical_facts:
                cic_score = critical_information_coverage(generated, critical_facts)

        # Answer Relevancy
        rel_score = None
        rel_questions: List[str] = []
        rel_sims: List[float] = []
        try:
            rel_score, rel_questions, rel_sims = calculate_answer_relevancy(
                pipe, query, generated
            )
        except Exception as e:
            print(f"[Answer Relevancy error] {e}")

        # Retrieval metrics
        mrr_score = None
        recall_score = None
        retrieved_ids: List[str] = []
        gold_ids: set = set()
        if ds_entry:
            keywords = ds_entry.get("mrr_relevant_keywords") or ds_entry.get(
                "critical_facts", []
            )

            chunks_mrr = pipe.searcher.search_similarity(query, limit=K_MRR)
            retrieved_ids = [chunk.id for chunk in chunks_mrr]

            if corpus:
                gold_ids = gold_ids_for_query(corpus, keywords, min_hits=MIN_HITS)
            else:
                # Fallback: gold sobre los recuperados
                gold_ids = {
                    chunk.id
                    for chunk in chunks_mrr
                    if len(
                        [
                            kw
                            for kw in keywords
                            if kw.lower() in chunk.text.lower()
                        ]
                    )
                    >= MIN_HITS
                }

            if gold_ids:
                rr, _ = reciprocal_rank(retrieved_ids, gold_ids)
                mrr_score = rr
                recall_score = recall_at_k(retrieved_ids, gold_ids, k=K_RECALL)

        elapsed = time.perf_counter() - t0

        result = {
            "id": idx,
            "fila_excel": row["fila"],
            "query": query,
            "expected_answer": expected,
            "generated_answer": generated,
            "tiempo_segundos": round(elapsed, 2),
            "sts": sts_score,
            "critical_info_coverage": cic_score,
            "critical_facts": ds_entry.get("critical_facts", []) if ds_entry else [],
            "answer_relevancy": rel_score,
            "preguntas_inversas": rel_questions,
            "similitudes_inversas": [round(s, 4) for s in rel_sims],
            "mrr": mrr_score,
            "recall_at_5": recall_score,
            "gold_keywords": ds_entry.get("mrr_relevant_keywords")
            or (ds_entry.get("critical_facts", []) if ds_entry else []),
            "gold_chunk_ids": sorted(gold_ids),
            "retrieved_chunk_ids": retrieved_ids,
        }
        results.append(result)

        # Acumular promedios
        for key in totals:
            val = result.get(key)
            if val is not None:
                totals[key] += val
                counts[key] += 1

        def fmt(value):
            return f"{value:.3f}" if value is not None else "N/A"

        print(
            f"{idx:<4} | {query[:44]:<45} | "
            f"{fmt(sts_score):>6} | {fmt(cic_score):>5} | "
            f"{fmt(rel_score):>5} | {fmt(mrr_score):>5} | {fmt(recall_score):>5}"
        )

    await pipe.on_shutdown()

    # 5. Promedios
    averages = {
        key: (totals[key] / counts[key] if counts[key] > 0 else None)
        for key in totals
    }

    print("\n" + "=" * 80)
    print(" PROMEDIOS GLOBALES")
    print("=" * 80)
    for key, avg in averages.items():
        label = {
            "sts": "STS",
            "critical_info_coverage": "Critical Information Coverage",
            "answer_relevancy": "Answer Relevancy",
            "mrr": "MRR@10",
            "recall_at_5": "Recall@5",
        }[key]
        if avg is not None:
            print(f"  {label:<35}: {avg:.4f}")
        else:
            print(f"  {label:<35}: N/A")
    print("=" * 80)

    # 6. Guardar JSON
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_model = model_name.replace(":", "_")
    json_path = os.path.join(
        OUTPUT_DIR, f"evaluacion_excel_completa_{safe_model}_{timestamp}.json"
    )

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "modelo": model_name,
                "metricas": ["sts", "critical_info_coverage", "answer_relevancy", "mrr", "recall_at_5"],
                "promedios_globales": averages,
                "total_preguntas": len(results),
                "resultados_detallados": results,
            },
            f,
            indent=4,
            ensure_ascii=False,
        )
    print(f"\nResultados guardados en: {json_path}")

    # 7. Guardar Excel con columnas nuevas
    excel_stem = Path(excel_path).stem
    excel_out = os.path.join(OUTPUT_DIR, f"{excel_stem}_completo_{safe_model}_{timestamp}.xlsx")
    guardar_excel_con_metricas(excel_path, excel_out, results)
    print(f"Excel con métricas guardado en: {excel_out}")


def guardar_excel_con_metricas(
    ruta_entrada: str, ruta_salida: str, results: List[Dict[str, Any]]
):
    """Copia el Excel original y agrega columnas con métricas y respuesta generada."""
    wb = openpyxl.load_workbook(ruta_entrada)
    ws = wb[SHEET_NAME]

    n_cols = len(next(ws.iter_rows(min_row=1, max_row=1)))
    col_respuesta = n_cols + 1
    col_sts = n_cols + 2
    col_cic = n_cols + 3
    col_rel = n_cols + 4
    col_mrr = n_cols + 5
    col_recall = n_cols + 6

    ws.cell(row=1, column=col_respuesta, value="respuesta_rag")
    ws.cell(row=1, column=col_sts, value="sts")
    ws.cell(row=1, column=col_cic, value="critical_info_coverage")
    ws.cell(row=1, column=col_rel, value="answer_relevancy")
    ws.cell(row=1, column=col_mrr, value="mrr_at_10")
    ws.cell(row=1, column=col_recall, value="recall_at_5")

    for r in results:
        fila = r["fila_excel"]
        ws.cell(row=fila, column=col_respuesta, value=r["generated_answer"])
        ws.cell(row=fila, column=col_sts, value=r["sts"])
        ws.cell(row=fila, column=col_cic, value=r["critical_info_coverage"])
        ws.cell(row=fila, column=col_rel, value=r["answer_relevancy"])
        ws.cell(row=fila, column=col_mrr, value=r["mrr"])
        ws.cell(row=fila, column=col_recall, value=r["recall_at_5"])

    wb.save(ruta_salida)
    wb.close()


if __name__ == "__main__":
    asyncio.run(main())
