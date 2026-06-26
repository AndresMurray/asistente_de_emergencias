"""
Ejecuta preguntas de un Excel contra el RAG vía OpenWebUI, calcula STS
contra las respuestas esperadas y guarda todo en un Excel de salida.

Uso:
    python rag_desde_excel.py "DataSet Test.xlsx" --apikey "sk-..." --modelo "main"
"""

import sys
import math
import time
import json
import argparse
from pathlib import Path

try:
    import openpyxl
    import requests
except ImportError as e:
    print(f"ERROR: falta dependencia — {e}. Instalá con: pip install openpyxl requests")
    sys.exit(1)

OPENWEBUI_URL = "http://localhost:8180"
OLLAMA_URL = "http://localhost:11434"
EMBED_MODEL = "paraphrase-multilingual"


# ── STS ──────────────────────────────────────────────────────────────────────

def get_embedding(text: str) -> list[float]:
    response = requests.post(
        f"{OLLAMA_URL}/api/embed",
        json={"model": EMBED_MODEL, "input": text},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["embeddings"][0]


def cosine_similarity(a: list[float], b: list[float]) -> float:
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x ** 2 for x in a))
    norm_b = math.sqrt(sum(x ** 2 for x in b))
    return dot / (norm_a * norm_b)


def sts(texto_a: str, texto_b: str) -> float:
    return cosine_similarity(get_embedding(texto_a), get_embedding(texto_b))


# ── Excel ─────────────────────────────────────────────────────────────────────

def leer_dataset(
    ruta_excel: str, col_pregunta: str, col_esperada: str, nombre_hoja: str | None
) -> list[tuple[int, str, str]]:
    """Retorna lista de (num_fila, pregunta, respuesta_esperada)."""
    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
    hoja = wb[nombre_hoja] if nombre_hoja else wb.active

    filas = hoja.iter_rows(values_only=True)
    encabezados = [str(c).strip() if c is not None else "" for c in next(filas)]

    def buscar_col(nombre: str) -> int:
        nombre_lower = nombre.lower()
        try:
            return next(i for i, h in enumerate(encabezados) if h.lower() == nombre_lower)
        except StopIteration:
            print(f"ERROR: no encontré la columna '{nombre}'.")
            print(f"  Columnas disponibles: {[h for h in encabezados if h]}")
            sys.exit(1)

    idx_pregunta = buscar_col(col_pregunta)
    idx_esperada = buscar_col(col_esperada)

    filas_data = []
    for num_fila, fila in enumerate(filas, start=2):
        if idx_pregunta >= len(fila):
            continue
        pregunta = fila[idx_pregunta]
        if pregunta is None or not str(pregunta).strip():
            continue
        esperada = fila[idx_esperada] if idx_esperada < len(fila) else None
        filas_data.append((num_fila, str(pregunta).strip(), str(esperada).strip() if esperada else ""))

    wb.close()
    return filas_data


def guardar_resultados(
    ruta_entrada: str,
    ruta_salida: str,
    nombre_hoja: str | None,
    resultados: list[tuple[int, str, str, float]],
) -> None:
    """Agrega columnas respuesta_rag y sts_score al Excel original."""
    wb = openpyxl.load_workbook(ruta_entrada)
    hoja = wb[nombre_hoja] if nombre_hoja else wb.active

    n_cols = len(next(hoja.iter_rows(min_row=1, max_row=1)))
    col_rag = n_cols + 1
    col_sts = n_cols + 2
    hoja.cell(row=1, column=col_rag, value="respuesta_rag")
    hoja.cell(row=1, column=col_sts, value="sts_score")

    for num_fila, _, respuesta, score in resultados:
        hoja.cell(row=num_fila, column=col_rag, value=respuesta)
        hoja.cell(row=num_fila, column=col_sts, value=round(score, 4) if score >= 0 else "ERROR")

    wb.save(ruta_salida)
    wb.close()


# ── RAG ───────────────────────────────────────────────────────────────────────

def preguntar_al_rag(pregunta: str, modelo: str, api_key: str) -> str:
    response = requests.post(
        f"{OPENWEBUI_URL}/api/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": modelo,
            "messages": [{"role": "user", "content": pregunta}],
            "stream": True,
        },
        stream=True,
        timeout=120,
    )
    response.raise_for_status()

    tokens = []
    for line in response.iter_lines():
        if not line:
            continue
        text = line.decode("utf-8") if isinstance(line, bytes) else line
        if not text.startswith("data:"):
            continue
        payload = text[len("data:"):].strip()
        if payload == "[DONE]":
            break
        data = json.loads(payload)
        content = data.get("choices", [{}])[0].get("delta", {}).get("content", "")
        if content:
            tokens.append(content)

    return "".join(tokens).strip()


def listar_modelos(api_key: str) -> None:
    response = requests.get(
        f"{OPENWEBUI_URL}/api/models",
        headers={"Authorization": f"Bearer {api_key}"},
        timeout=10,
    )
    response.raise_for_status()
    for m in response.json().get("data", []):
        print(f"  - {m['id']}")
    sys.exit(0)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", help="Ruta al archivo Excel")
    parser.add_argument("--columna", default="Query", help="Columna de preguntas (default: Query)")
    parser.add_argument("--esperada", default="Respuesta Esperada", help="Columna de respuestas esperadas (default: 'Respuesta Esperada')")
    parser.add_argument("--hoja", default=None)
    parser.add_argument("--salida", default=None)
    parser.add_argument("--apikey", required=True)
    parser.add_argument("--modelo", default=None)
    parser.add_argument("--listar-modelos", action="store_true")
    args = parser.parse_args()

    if args.listar_modelos:
        listar_modelos(args.apikey)

    if not args.modelo:
        print("ERROR: especificá --modelo. Usá --listar-modelos para ver los disponibles.")
        sys.exit(1)

    if not Path(args.excel).exists():
        print(f"ERROR: no existe '{args.excel}'")
        sys.exit(1)

    ruta_salida = args.salida or str(Path(args.excel).with_stem(Path(args.excel).stem + "_con_respuestas"))

    print(f"\n{'═'*60}")
    print(f"  RAG desde Excel  →  OpenWebUI + STS")
    print(f"{'═'*60}")
    print(f"  Archivo:  {args.excel}")
    print(f"  Modelo:   {args.modelo}")
    print(f"  Salida:   {ruta_salida}")
    print(f"{'═'*60}\n")

    dataset = leer_dataset(args.excel, args.columna, args.esperada, args.hoja)
    if not dataset:
        print("No encontré preguntas.")
        sys.exit(1)

    print(f"  {len(dataset)} pregunta(s) encontradas.\n")

    # (num_fila, pregunta, respuesta_rag, sts_score)
    resultados: list[tuple[int, str, str, float]] = []

    for idx, (num_fila, pregunta, esperada) in enumerate(dataset, start=1):
        print(f"{'─'*60}")
        print(f"  [{idx}/{len(dataset)}] {pregunta}")
        print(f"{'─'*60}")

        t0 = time.perf_counter()
        try:
            respuesta = preguntar_al_rag(pregunta, args.modelo, args.apikey)
        except Exception as e:
            respuesta = f"ERROR: {e}"

        try:
            score = sts(respuesta, esperada) if esperada and not respuesta.startswith("ERROR") else -1.0
        except Exception as e:
            print(f"  [STS] Error calculando similitud: {e}")
            score = -1.0

        elapsed = time.perf_counter() - t0

        print(f"\n  RESPUESTA ({elapsed:.1f}s):\n")
        for linea in respuesta.split("\n"):
            print(f"    {linea}")

        score_str = f"{score:.4f} ({score*100:.1f}%)" if score >= 0 else "N/A"
        print(f"\n  STS score: {score_str}\n")

        resultados.append((num_fila, pregunta, respuesta, score))

    scores_validos = [r[3] for r in resultados if r[3] >= 0]
    if scores_validos:
        promedio = sum(scores_validos) / len(scores_validos)
        print(f"{'═'*60}")
        print(f"  Promedio STS: {promedio:.4f} ({promedio*100:.1f}%)")

    print(f"{'═'*60}")
    print(f"  Guardando en '{ruta_salida}'...")
    guardar_resultados(args.excel, ruta_salida, args.hoja, resultados)
    print(f"  Listo. {len(resultados)} fila(s) guardadas.")
    print(f"{'═'*60}\n")


if __name__ == "__main__":
    main()
